import os
import sys
import xlwt
import xlrd
from xlutils.copy import copy
import re
from vhd_tb.cutils import get_vector_conf
class xls(object):
    def __init__(self, file_name):
        self.file_name = file_name

    def get_format(self):
        return 'xls'

    def create(self,tables, pdic):
        doc = xlwt.Workbook()
        for table in tables:
            sheet = doc.add_sheet(table)
            i = 0
            for port in tables[table]:
                sheet.write(0,i,port+'['+`pdic[port][2]`+']')
                i += 1
        doc.save(self.file_name)

    def add_sheet(self,prefix,tables,pdic):
        wb = xlrd.open_workbook(self.file_name,formatting_info=True)
        doc = xlwt.Workbook()
        for table in tables:
            sheet = doc.add_sheet(table)
            i=0
            for port in tables[table]:
                sheet.write(0,i,port+'['+`pdic[port][2]`+']')
                i += 1
        for sheet_name in wb.sheet_names():
            sheet = wb.sheet_by_name(sheet_name)
            if len(sheet_name.split('-')) == 2:
                new_name = sheet_name
            else:
                new_name = '%s-%s'%(prefix,sheet_name)
            copy_sheet = doc.add_sheet(new_name)
            for i in range(sheet.nrows):
                for j in range(sheet.ncols):
                    copy_sheet.write(i,j,sheet.cell(i,j).value)
        doc.save(self.file_name)
        

    def get_table(self,tables,pdic, tb_name):
        doc = xlrd.open_workbook(self.file_name, formatting_info=True)
        dst_dir = os.path.dirname(os.path.abspath(self.file_name))
        for table in tables:
            sheet = doc.sheet_by_name(table)
            ports = tables[table]
            vector_mult = get_vector_conf(ports,pdic)[0]
            file_name = '%s/%s_%s_%s.tbl'%(dst_dir,tb_name,'xls',table)
            f = open(file_name,'w')
            for row in range(1, sheet.nrows):
                data = 0
                for col in range(sheet.ncols):
                    intr = sheet.cell(row,col).value
                    data+=int(intr)*vector_mult[col]
                f.write(`data`+'\n')

            f.close()


from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter

class xlsx(object):
    def __init__(self, file_name):
        self.file_name = file_name

    def get_format(self):
        return 'xlsx'

    def create(self, tables, pdic):
        doc = Workbook()
        ew = ExcelWriter(workbook = doc)
        for table in tables:
            sheet = doc.worksheets[0]
            sheet.title = table
            i = 0
            for port in tables[table]:
                c = sheet.cell(row=0,column=i)

                c.value= port+'['+`pdic[port][2]`+']'
                i += 1
        try:
            ew.save(filename = self.file_name)
            return True
        except:
            return False

    def add_sheet(self,perfix,tables,pdic):
        doc = load_workbook(self.file_name)
        for sheet_name in doc.get_sheet_names():
            sheet = doc.get_sheet_by_name(sheet_name)
            if len(sheet_name.split('-')) == 2:
                new_name = sheet_name
            else:
                new_name = '%s-%s'%(prefix,sheet.title)
            sheet.title = new_name

        for table in tables:
            sheet = doc.create_sheet(0)
            sheet.title = table
            i = 0
            for port in tables[table]:
                c = sheet.cell(row=0,column=i)

                c.value= port+'['+`pdic[port][2]`+']'
                i += 1
        doc.save(self.file_name)

    def get_table(self, tables, pdic, tb_name):
        doc = load_workbook(self.file_name)
        dst_dir = os.path.dirname(os.path.abspath(self.file_name))
        for table in tables:
            sheet = doc.get_sheet_by_name(table)
            ports = tables[table]
            vector_mult = get_vector_conf(ports, pdic)[0]
            file_name = '%s/%s_%s_%s.tbl'%(dst_dir,tb_name,'xlsx',table)
            f = open(file_name,'w')
            for row in range(1,sheet.get_highest_row()):
                data = 0
                for col in range(sheet.get_highest_column()):
                    data += sheet.cell(row=row, column=col).value
                f.write(`data`+'\n')

            f.close()


import ooolib

class ods(object):
    def __init__(self,file_name):
        self.file_name = file_name

    def get_format(self):
        return 'ods'

    def create(self, tables, pdic):
        doc = ooolib.Calc()
        sheets  = []
        for table in tables:
            sheet = ooolib.CalcSheet(table)
            i=1
            for port in tables[table]:
                sheet.set_sheet_value((i,1),"string",port+'['+`pdic[port][2]`+']')
                i += 1
            sheets.append(sheet)
        doc.sheets = sheets
        doc.save(self.file_name)


    def add_sheet(self,prefix, tables, pdic):
        doc = ooolib.Calc(opendoc=self.file_name)
        sheets = doc.sheets
        for sheet in sheets:
            sheet_name = sheet.get_name()
            if len(sheet_name.split('-')) == 2:
                new_name = sheet_name
            else:
                new_name = '%s-%s'%(prefix,sheet_name)
            sheet.set_name(new_name)
        doc.sheets = sheets
        for table in tables:
            doc.new_sheet(table)
            i=1
            for port in tables[table]:
                doc.set_cell_value(i,1,"string",port+'['+`pdic[port][2]`+']')
                i += 1
        doc.save(self.file_name)

    def get_table(self, tables, pdic, tb_name):
        doc = ooolib.Calc(opendoc = self.file_name)
        dst_dir = os.path.dirname(os.path.abspath(self.file_name))
        sheets = {}
        for sheet in doc.sheets:
            sheets[sheet.get_name()] = sheet
        for table in tables:
            sheet = sheets[table]
            ports = tables[table]
            vector_mult = get_vector_conf(ports,pdic)[0]
            file_name = '%s/%s_%s_%s.tbl'%(dst_dir,tb_name,'ods',table)
            f = open(file_name,'w')
            (cols, rows) = sheet.get_sheet_dimensions()
            for row in range(2, rows+1):
                data = 0
                for col in range(1, cols+1):
                    intr = sheet.get_sheet_value(col, row)['value']
                    data += int(intr[1])*vector_mult[col-1]
                f.write(`data`+"\n")
            
            f.close()
